"""
Lookup Tables for Logic Weaver.

Provides reference data management:
- Table definition and storage
- Query operations with filtering
- Import/Export (CSV, Excel, JSON)
- Caching for performance

Lookup tables enable business users to manage reference data
that can be used in mappings and business rules.
"""

from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from enum import Enum
from typing import Any, Callable, Dict, List, Optional, Set, Tuple, Union
import json
import csv
import io
import hashlib
from datetime import datetime, timedelta


# =============================================================================
# Column Types
# =============================================================================


class ColumnType(Enum):
    """Data types for lookup table columns."""
    STRING = "string"
    NUMBER = "number"
    INTEGER = "integer"
    BOOLEAN = "boolean"
    DATE = "date"
    DATETIME = "datetime"
    JSON = "json"


@dataclass
class LookupColumn:
    """Definition of a lookup table column."""

    name: str
    type: ColumnType
    description: Optional[str] = None
    required: bool = False
    unique: bool = False
    default_value: Any = None
    primary_key: bool = False

    # Validation
    min_length: Optional[int] = None
    max_length: Optional[int] = None
    min_value: Optional[float] = None
    max_value: Optional[float] = None
    allowed_values: Optional[List[Any]] = None
    pattern: Optional[str] = None

    def validate(self, value: Any) -> List[str]:
        """Validate a value against column definition."""
        errors = []

        # Required check
        if self.required and (value is None or value == ""):
            errors.append(f"{self.name}: Value is required")
            return errors

        if value is None:
            return errors

        # Type check
        if self.type == ColumnType.NUMBER:
            if not isinstance(value, (int, float)):
                try:
                    float(value)
                except (ValueError, TypeError):
                    errors.append(f"{self.name}: Must be a number")
                    return errors

        elif self.type == ColumnType.INTEGER:
            if not isinstance(value, int):
                try:
                    int(value)
                except (ValueError, TypeError):
                    errors.append(f"{self.name}: Must be an integer")
                    return errors

        elif self.type == ColumnType.BOOLEAN:
            if not isinstance(value, bool):
                if str(value).lower() not in ("true", "false", "1", "0", "yes", "no"):
                    errors.append(f"{self.name}: Must be a boolean")

        # Length validation
        if isinstance(value, str):
            if self.min_length and len(value) < self.min_length:
                errors.append(f"{self.name}: Minimum length is {self.min_length}")
            if self.max_length and len(value) > self.max_length:
                errors.append(f"{self.name}: Maximum length is {self.max_length}")

        # Range validation
        if isinstance(value, (int, float)):
            if self.min_value is not None and value < self.min_value:
                errors.append(f"{self.name}: Minimum value is {self.min_value}")
            if self.max_value is not None and value > self.max_value:
                errors.append(f"{self.name}: Maximum value is {self.max_value}")

        # Allowed values
        if self.allowed_values and value not in self.allowed_values:
            errors.append(f"{self.name}: Must be one of: {', '.join(map(str, self.allowed_values))}")

        # Pattern
        if self.pattern:
            import re
            if not re.match(self.pattern, str(value)):
                errors.append(f"{self.name}: Does not match required pattern")

        return errors

    def coerce(self, value: Any) -> Any:
        """Coerce value to column type."""
        if value is None:
            return self.default_value

        try:
            if self.type == ColumnType.STRING:
                return str(value)
            elif self.type == ColumnType.NUMBER:
                return float(value)
            elif self.type == ColumnType.INTEGER:
                return int(float(value))
            elif self.type == ColumnType.BOOLEAN:
                if isinstance(value, bool):
                    return value
                return str(value).lower() in ("true", "yes", "1", "on")
            elif self.type == ColumnType.DATE:
                if isinstance(value, datetime):
                    return value.date()
                return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
            elif self.type == ColumnType.DATETIME:
                if isinstance(value, datetime):
                    return value
                return datetime.fromisoformat(str(value))
            elif self.type == ColumnType.JSON:
                if isinstance(value, (dict, list)):
                    return value
                return json.loads(value)
        except (ValueError, TypeError, json.JSONDecodeError):
            return value

        return value


# =============================================================================
# Lookup Row
# =============================================================================


@dataclass
class LookupRow:
    """A row in a lookup table."""

    data: Dict[str, Any]
    row_id: Optional[str] = None
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None
    created_by: Optional[str] = None
    version: int = 1

    def __post_init__(self):
        if not self.row_id:
            # Generate row ID from data hash
            self.row_id = hashlib.md5(
                json.dumps(self.data, sort_keys=True, default=str).encode()
            ).hexdigest()[:16]
        if not self.created_at:
            self.created_at = datetime.utcnow()
        if not self.updated_at:
            self.updated_at = self.created_at

    def get(self, column: str, default: Any = None) -> Any:
        """Get column value."""
        return self.data.get(column, default)

    def matches(self, filters: Dict[str, Any]) -> bool:
        """Check if row matches filters."""
        for key, value in filters.items():
            row_value = self.data.get(key)

            # Handle operators in filter
            if isinstance(value, dict):
                op = value.get("op", "eq")
                val = value.get("value")

                if op == "eq" and row_value != val:
                    return False
                elif op == "ne" and row_value == val:
                    return False
                elif op == "gt" and not (row_value is not None and row_value > val):
                    return False
                elif op == "gte" and not (row_value is not None and row_value >= val):
                    return False
                elif op == "lt" and not (row_value is not None and row_value < val):
                    return False
                elif op == "lte" and not (row_value is not None and row_value <= val):
                    return False
                elif op == "contains" and not (row_value and val in str(row_value)):
                    return False
                elif op == "starts_with" and not (row_value and str(row_value).startswith(str(val))):
                    return False
                elif op == "ends_with" and not (row_value and str(row_value).endswith(str(val))):
                    return False
                elif op == "in" and row_value not in val:
                    return False
                elif op == "not_in" and row_value in val:
                    return False
            else:
                if row_value != value:
                    return False

        return True


# =============================================================================
# Lookup Table
# =============================================================================


@dataclass
class LookupTable:
    """A lookup table for reference data."""

    id: str
    name: str
    columns: List[LookupColumn]
    description: Optional[str] = None
    tenant_id: Optional[str] = None
    rows: List[LookupRow] = field(default_factory=list)

    # Metadata
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None
    created_by: Optional[str] = None
    version: int = 1

    # Settings
    max_rows: int = 100000
    enable_history: bool = False

    def __post_init__(self):
        if not self.created_at:
            self.created_at = datetime.utcnow()
        if not self.updated_at:
            self.updated_at = self.created_at

    def get_primary_key_column(self) -> Optional[LookupColumn]:
        """Get the primary key column."""
        for col in self.columns:
            if col.primary_key:
                return col
        return self.columns[0] if self.columns else None

    def get_column(self, name: str) -> Optional[LookupColumn]:
        """Get column by name."""
        for col in self.columns:
            if col.name == name:
                return col
        return None

    def add_row(
        self,
        data: Dict[str, Any],
        validate: bool = True,
        coerce: bool = True
    ) -> Tuple[bool, Union[LookupRow, List[str]]]:
        """
        Add a row to the table.

        Returns (success, row_or_errors)
        """
        if len(self.rows) >= self.max_rows:
            return False, ["Table has reached maximum row limit"]

        # Coerce values
        if coerce:
            coerced_data = {}
            for col in self.columns:
                value = data.get(col.name)
                coerced_data[col.name] = col.coerce(value)
            data = coerced_data

        # Validate
        if validate:
            errors = self.validate_row(data)
            if errors:
                return False, errors

        # Check uniqueness
        pk_col = self.get_primary_key_column()
        if pk_col and pk_col.unique:
            pk_value = data.get(pk_col.name)
            for existing in self.rows:
                if existing.get(pk_col.name) == pk_value:
                    return False, [f"Duplicate primary key: {pk_value}"]

        row = LookupRow(data=data)
        self.rows.append(row)
        self.updated_at = datetime.utcnow()
        self.version += 1

        return True, row

    def update_row(
        self,
        row_id: str,
        data: Dict[str, Any],
        validate: bool = True,
        coerce: bool = True
    ) -> Tuple[bool, Union[LookupRow, List[str]]]:
        """Update a row by ID."""
        for i, row in enumerate(self.rows):
            if row.row_id == row_id:
                # Merge with existing data
                new_data = {**row.data, **data}

                if coerce:
                    coerced_data = {}
                    for col in self.columns:
                        value = new_data.get(col.name)
                        coerced_data[col.name] = col.coerce(value)
                    new_data = coerced_data

                if validate:
                    errors = self.validate_row(new_data)
                    if errors:
                        return False, errors

                row.data = new_data
                row.updated_at = datetime.utcnow()
                row.version += 1
                self.updated_at = datetime.utcnow()
                self.version += 1

                return True, row

        return False, ["Row not found"]

    def delete_row(self, row_id: str) -> bool:
        """Delete a row by ID."""
        for i, row in enumerate(self.rows):
            if row.row_id == row_id:
                self.rows.pop(i)
                self.updated_at = datetime.utcnow()
                self.version += 1
                return True
        return False

    def validate_row(self, data: Dict[str, Any]) -> List[str]:
        """Validate row data against column definitions."""
        errors = []
        for col in self.columns:
            value = data.get(col.name)
            col_errors = col.validate(value)
            errors.extend(col_errors)
        return errors

    def lookup(
        self,
        key_column: str,
        key_value: Any,
        value_column: Optional[str] = None
    ) -> Any:
        """
        Quick lookup by key.

        Args:
            key_column: Column to search
            key_value: Value to find
            value_column: Column to return (None for full row)

        Returns:
            Found value, row, or None
        """
        for row in self.rows:
            if row.get(key_column) == key_value:
                if value_column:
                    return row.get(value_column)
                return row.data
        return None

    def query(
        self,
        filters: Optional[Dict[str, Any]] = None,
        sort_by: Optional[str] = None,
        sort_desc: bool = False,
        limit: Optional[int] = None,
        offset: int = 0
    ) -> List[LookupRow]:
        """
        Query rows with filtering and pagination.

        Args:
            filters: Column filters
            sort_by: Column to sort by
            sort_desc: Sort descending
            limit: Maximum rows to return
            offset: Rows to skip

        Returns:
            Matching rows
        """
        results = self.rows

        # Filter
        if filters:
            results = [r for r in results if r.matches(filters)]

        # Sort
        if sort_by:
            results = sorted(
                results,
                key=lambda r: r.get(sort_by) or "",
                reverse=sort_desc
            )

        # Paginate
        if offset:
            results = results[offset:]
        if limit:
            results = results[:limit]

        return results

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            "id": self.id,
            "name": self.name,
            "description": self.description,
            "columns": [
                {
                    "name": col.name,
                    "type": col.type.value,
                    "description": col.description,
                    "required": col.required,
                    "unique": col.unique,
                    "primaryKey": col.primary_key,
                    "defaultValue": col.default_value
                }
                for col in self.columns
            ],
            "rows": [
                {
                    "id": row.row_id,
                    "data": row.data,
                    "createdAt": row.created_at.isoformat() if row.created_at else None,
                    "updatedAt": row.updated_at.isoformat() if row.updated_at else None
                }
                for row in self.rows
            ],
            "version": self.version,
            "createdAt": self.created_at.isoformat() if self.created_at else None,
            "updatedAt": self.updated_at.isoformat() if self.updated_at else None
        }


# =============================================================================
# Query Operations
# =============================================================================


@dataclass
class LookupQuery:
    """Query definition for lookup tables."""

    table_id: str
    filters: Dict[str, Any] = field(default_factory=dict)
    select_columns: Optional[List[str]] = None
    sort_by: Optional[str] = None
    sort_desc: bool = False
    limit: Optional[int] = None
    offset: int = 0

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            "tableId": self.table_id,
            "filters": self.filters,
            "selectColumns": self.select_columns,
            "sortBy": self.sort_by,
            "sortDesc": self.sort_desc,
            "limit": self.limit,
            "offset": self.offset
        }


@dataclass
class LookupResult:
    """Result of a lookup query."""

    success: bool
    rows: List[Dict[str, Any]] = field(default_factory=list)
    total_count: int = 0
    error: Optional[str] = None
    execution_time_ms: float = 0

    @classmethod
    def ok(cls, rows: List[LookupRow], total_count: int = 0) -> "LookupResult":
        """Create successful result."""
        return cls(
            success=True,
            rows=[r.data for r in rows],
            total_count=total_count or len(rows)
        )

    @classmethod
    def fail(cls, error: str) -> "LookupResult":
        """Create failed result."""
        return cls(success=False, error=error)


# =============================================================================
# Lookup Table Manager
# =============================================================================


class LookupTableManager:
    """
    Manages lookup tables.

    In production, this would be backed by a database.
    This implementation provides in-memory storage.
    """

    def __init__(self):
        """Initialize manager."""
        self._tables: Dict[str, LookupTable] = {}
        self._cache = LookupCache()

    def create_table(
        self,
        table_id: str,
        name: str,
        columns: List[LookupColumn],
        description: Optional[str] = None,
        tenant_id: Optional[str] = None
    ) -> LookupTable:
        """Create a new lookup table."""
        table = LookupTable(
            id=table_id,
            name=name,
            columns=columns,
            description=description,
            tenant_id=tenant_id
        )
        self._tables[table_id] = table
        return table

    def get_table(self, table_id: str) -> Optional[LookupTable]:
        """Get a table by ID."""
        return self._tables.get(table_id)

    def delete_table(self, table_id: str) -> bool:
        """Delete a table."""
        if table_id in self._tables:
            del self._tables[table_id]
            self._cache.invalidate(table_id)
            return True
        return False

    def list_tables(
        self,
        tenant_id: Optional[str] = None
    ) -> List[LookupTable]:
        """List all tables, optionally filtered by tenant."""
        tables = list(self._tables.values())
        if tenant_id:
            tables = [t for t in tables if t.tenant_id == tenant_id]
        return tables

    def lookup(
        self,
        table_id: str,
        key_column: str,
        key_value: Any,
        value_column: Optional[str] = None,
        use_cache: bool = True
    ) -> Any:
        """
        Quick lookup with optional caching.

        Args:
            table_id: Table ID
            key_column: Column to search
            key_value: Value to find
            value_column: Column to return
            use_cache: Whether to use cache

        Returns:
            Found value or None
        """
        cache_key = f"{table_id}:{key_column}:{key_value}:{value_column}"

        if use_cache:
            cached = self._cache.get(cache_key)
            if cached is not None:
                return cached

        table = self.get_table(table_id)
        if not table:
            return None

        result = table.lookup(key_column, key_value, value_column)

        if use_cache and result is not None:
            self._cache.set(cache_key, result)

        return result

    def query(self, query: LookupQuery) -> LookupResult:
        """Execute a lookup query."""
        import time
        start = time.time()

        table = self.get_table(query.table_id)
        if not table:
            return LookupResult.fail(f"Table not found: {query.table_id}")

        try:
            rows = table.query(
                filters=query.filters,
                sort_by=query.sort_by,
                sort_desc=query.sort_desc,
                limit=query.limit,
                offset=query.offset
            )

            # Apply column selection
            if query.select_columns:
                for row in rows:
                    row.data = {k: v for k, v in row.data.items() if k in query.select_columns}

            result = LookupResult.ok(rows, len(table.rows))
            result.execution_time_ms = (time.time() - start) * 1000
            return result

        except Exception as e:
            return LookupResult.fail(str(e))

    def bulk_insert(
        self,
        table_id: str,
        rows: List[Dict[str, Any]],
        validate: bool = True,
        coerce: bool = True
    ) -> Tuple[int, List[str]]:
        """
        Bulk insert rows.

        Returns (inserted_count, errors)
        """
        table = self.get_table(table_id)
        if not table:
            return 0, ["Table not found"]

        inserted = 0
        errors = []

        for i, row_data in enumerate(rows):
            success, result = table.add_row(row_data, validate=validate, coerce=coerce)
            if success:
                inserted += 1
            else:
                errors.append(f"Row {i}: {', '.join(result)}")

        self._cache.invalidate(table_id)
        return inserted, errors


# =============================================================================
# Cache
# =============================================================================


class LookupCache:
    """Simple in-memory cache for lookup results."""

    def __init__(self, ttl_seconds: int = 300, max_size: int = 10000):
        """Initialize cache."""
        self._cache: Dict[str, Tuple[Any, datetime]] = {}
        self._ttl = timedelta(seconds=ttl_seconds)
        self._max_size = max_size

    def get(self, key: str) -> Optional[Any]:
        """Get cached value."""
        if key in self._cache:
            value, expires = self._cache[key]
            if datetime.utcnow() < expires:
                return value
            else:
                del self._cache[key]
        return None

    def set(self, key: str, value: Any) -> None:
        """Set cached value."""
        # Evict if at capacity
        if len(self._cache) >= self._max_size:
            # Remove oldest entries
            sorted_keys = sorted(
                self._cache.keys(),
                key=lambda k: self._cache[k][1]
            )
            for k in sorted_keys[:len(sorted_keys) // 4]:
                del self._cache[k]

        self._cache[key] = (value, datetime.utcnow() + self._ttl)

    def invalidate(self, prefix: str) -> None:
        """Invalidate all keys with prefix."""
        keys_to_delete = [k for k in self._cache if k.startswith(prefix)]
        for k in keys_to_delete:
            del self._cache[k]

    def clear(self) -> None:
        """Clear all cached values."""
        self._cache.clear()


# =============================================================================
# Import/Export
# =============================================================================


class TableImporter:
    """Imports data into lookup tables."""

    def import_csv(
        self,
        table: LookupTable,
        csv_data: str,
        has_header: bool = True,
        delimiter: str = ","
    ) -> Tuple[int, List[str]]:
        """
        Import from CSV.

        Returns (imported_count, errors)
        """
        reader = csv.reader(io.StringIO(csv_data), delimiter=delimiter)
        rows = list(reader)

        if not rows:
            return 0, ["No data to import"]

        # Get headers
        if has_header:
            headers = rows[0]
            data_rows = rows[1:]
        else:
            headers = [col.name for col in table.columns]
            data_rows = rows

        imported = 0
        errors = []

        for i, row in enumerate(data_rows):
            if len(row) != len(headers):
                errors.append(f"Row {i + 1}: Column count mismatch")
                continue

            row_data = dict(zip(headers, row))
            success, result = table.add_row(row_data, validate=True, coerce=True)

            if success:
                imported += 1
            else:
                errors.append(f"Row {i + 1}: {', '.join(result)}")

        return imported, errors

    def import_json(
        self,
        table: LookupTable,
        json_data: str
    ) -> Tuple[int, List[str]]:
        """
        Import from JSON.

        Returns (imported_count, errors)
        """
        try:
            data = json.loads(json_data)
        except json.JSONDecodeError as e:
            return 0, [f"Invalid JSON: {e}"]

        if not isinstance(data, list):
            data = [data]

        imported = 0
        errors = []

        for i, row_data in enumerate(data):
            if not isinstance(row_data, dict):
                errors.append(f"Row {i + 1}: Must be an object")
                continue

            success, result = table.add_row(row_data, validate=True, coerce=True)

            if success:
                imported += 1
            else:
                errors.append(f"Row {i + 1}: {', '.join(result)}")

        return imported, errors


class TableExporter:
    """Exports data from lookup tables."""

    def export_csv(
        self,
        table: LookupTable,
        include_header: bool = True,
        columns: Optional[List[str]] = None
    ) -> str:
        """Export to CSV."""
        output = io.StringIO()
        writer = csv.writer(output)

        # Determine columns
        if columns:
            col_names = columns
        else:
            col_names = [col.name for col in table.columns]

        # Write header
        if include_header:
            writer.writerow(col_names)

        # Write rows
        for row in table.rows:
            writer.writerow([row.get(col) for col in col_names])

        return output.getvalue()

    def export_json(
        self,
        table: LookupTable,
        columns: Optional[List[str]] = None,
        pretty: bool = True
    ) -> str:
        """Export to JSON."""
        data = []

        for row in table.rows:
            if columns:
                row_data = {k: v for k, v in row.data.items() if k in columns}
            else:
                row_data = row.data
            data.append(row_data)

        if pretty:
            return json.dumps(data, indent=2, default=str)
        return json.dumps(data, default=str)

    def export_excel(
        self,
        table: LookupTable,
        columns: Optional[List[str]] = None
    ) -> bytes:
        """
        Export to Excel format.

        Note: Requires openpyxl - production would include this.
        Returns CSV as fallback if openpyxl not available.
        """
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = table.name[:31]  # Excel sheet name limit

            # Determine columns
            col_names = columns or [col.name for col in table.columns]

            # Write header
            for i, name in enumerate(col_names, 1):
                ws.cell(row=1, column=i, value=name)

            # Write data
            for row_idx, row in enumerate(table.rows, 2):
                for col_idx, col_name in enumerate(col_names, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        except ImportError:
            # Fallback to CSV
            return self.export_csv(table, columns=columns).encode()


# =============================================================================
# Exports
# =============================================================================

__all__ = [
    # Table types
    "LookupTable",
    "LookupColumn",
    "ColumnType",
    "LookupRow",
    # Operations
    "LookupQuery",
    "LookupResult",
    # Manager
    "LookupTableManager",
    "LookupCache",
    # Import/Export
    "TableImporter",
    "TableExporter",
]
